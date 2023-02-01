import pandas as pd
from openpyxl import load_workbook
import yfinance as yf
import excel2img
import json
from datetime import datetime as dt
import os

def wig20_components():
    wig20 = pd.read_html('https://www.biznesradar.pl/gielda/indeks:WIG20')[0]
    wig20['Ticker'] = wig20.Profil.str[:3]
    wig20.drop(['Czas', '1r 3m'], axis=1, inplace=True)

    return wig20


def wig20_earnings_calendar():
    raporty_wig20 = pd.read_html(
        'https://strefainwestorow.pl/dane/raporty/lista-dat-publikacji-raportow-okresowych/wig20')[0]

    raporty_wig20.Ticker = raporty_wig20.Ticker.str.lstrip('#')

    dates = dict()
    for i, row in raporty_wig20.iterrows():
        _, ticker, _, data, raport = row

        dates.setdefault(data, {ticker: raport}).update({ticker: raport})

    with open('wig20_earnings_dates', 'w', encoding='utf-8') as f:
        json.dump(dates, f)

    return dates


def wig40_earnings_calendar():

    raporty_wig20 = pd.read_html(
        'https://strefainwestorow.pl/dane/raporty/lista-dat-publikacji-raportow-okresowych/mwig40')[0]

    raporty_wig20.Ticker = raporty_wig20.Ticker.str.lstrip('#')

    dates = dict()
    for i, row in raporty_wig20.iterrows():
        _, ticker, _, data, raport = row

        dates.setdefault(data, {ticker: raport}).update({ticker: raport})

    with open('wig20_earnings_dates', 'w', encoding='utf-8') as f:
        json.dump(dates, f)

    return dates


def prepare_template(ticker):

    ticker = 'PKN'

    full_ticker = ticker + '.WA'


    stonk = yf.Ticker(full_ticker)
    rev = stonk.get_rev_forecast()
    earn = stonk.get_earnings_forecast()

    wb = load_workbook(filename='szablon.xlsx')
    ws = wb.active
    ws['C3'] = full_ticker
    ws['C7'] = ticker
    noww = dt.now()
    noww = dt.strftime(noww, r'%Y/%m/%d %H:%M:%S')
    ws['C20'] = noww

    ws['D9'] = int(rev[rev.period == '0q'].avg)
    ws['F9'] = float(earn[earn.period == '0q'].avg)

    ws['D16'] = int(rev[rev.period == '0q'].avg)
    ws['F16'] = float(earn[earn.period == '0q'].avg)

    wb.save(f'{ticker}.xlsx')

    print(f'szablon dla {ticker} przygotowany')


def chart_cons_vs_actual(ticker):

    full_ticker = ticker + '.WA'
    stonk = yf.Ticker(full_ticker)

    incm = stonk.get_income_stmt(freq='quarterly')
    rev = incm.loc['TotalRevenue'][0]
    eps = incm.loc['DilutedEPS'][0]

    wb = load_workbook(filename=f'{ticker}.xlsx')
    ws = wb.active
    noww = dt.now()
    noww = dt.strftime(noww, r'%Y/%m/%d %H:%M:%S')
    ws['C20'] = noww

    ws['D10'] = int(rev)
    ws['F10'] = eps

    ws['D17'] = int(rev)
    ws['F17'] = eps

    wb.save(f'{ticker}.xlsx')
    print('nowe dane wstawione')


    excel2img.export_img(
        f'{ticker}.xlsx', f'{ticker}.png', '', 'Arkusz1!B2:G20')
    print(f'obrazek {ticker} gotowy')

    os.remove(f'{ticker}.xlsx')
    print('excel usuniety')


def companies_reporting(date: str):

    with open('wig20_earnings_dates', 'r', encoding='utf-8') as f:
        dates: dict = json.load(f)

    if date in dates:
        return dates[date]
    else:
        return None


if __name__ == '__main__':

    prepare_template('PKN')
    # chart_cons_vs_actual('PKN')

    pass
